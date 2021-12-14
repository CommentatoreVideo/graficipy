from bokeh.models.callbacks import Callback
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pickle as pickle
#apro il file excel precedentemente caricato
#Inserisci il nome visualizzato sopra, dopo il caricamento
# nomefilecaricato="owid-covid-data_sint_beta.xlsx"
nomefilecaricato="owid-covid-data_sint_beta3.xlsx"
print("Inizio a caricare il file")
excel_document = openpyxl.load_workbook(nomefilecaricato)
print("File caricato")
#type(excel_document)
#excel_document.get_sheet_names()
#carico il Foglio1
print("Inizio a caricare il dataset")
sheet = excel_document.get_sheet_by_name('Sheet1')
print("Dataset caricato")
numRecord=sheet.max_row-1
print(numRecord)
daEvitare=["Asia","World"]
valide=["Asia","World"]
numCol=sheet.max_column
def getData(sheet,riga):
  return sheet.cell(row=riga,column=4).value
def getLocation(sheet,riga):
  return sheet.cell(row=riga,column=3).value
def getTotalDeath(sheet,riga):
  return sheet.cell(row=riga,column=5).value
def nazioneValida(nazione):
  evitare=False
  tutte=True
  if tutte==True:
    return True
  if evitare:
    return nazione not in daEvitare
  return nazione in valide
# Prendo l'elenco delle nazioni
print("Prendo le nazioni")
nazioni=[]
for i in range(2,numRecord+1):
  if nazioneValida(getLocation(sheet,i)):
    nazioni.append(getLocation(sheet,i))
nazioni=list(set(nazioni))
print("Nazioni prese")
nazioni.sort()
# print(nazioni)
# Prendo le date
date=[]
assi_y={}
print("Prendo date e casi totali")
for i in range(2,numRecord+1):
  date.append(getData(sheet,i))
  # Prendo i valori dei casi totali
  location=getLocation(sheet,i)
  if location not in nazioni:
    continue
  if assi_y.get(location)==None:
    assi_y[location]=[]
  assi_y[location].append([getTotalDeath(sheet,i),getData(sheet,i)])
print("Nazioni e casi totali presi")
from bokeh.plotting import figure,show
from bokeh.models import HoverTool, Callback
from bokeh.core.properties import Instance
date=list(set(date))
date.sort()
# Aggiungo date mancanti
print("Aggiungo date mancanti")
for nazione in nazioni:
  datePresenti=[]
  try:
    # Prendo i valori dei casi totali
    dati=assi_y[nazione]
    for d in dati:
      datePresenti.append(d[1])
    for data in date:
      if data not in datePresenti:
        assi_y[nazione].append([0,data])
    assi_y[nazione].sort(key=lambda x: x[1])
  except KeyError as e:
    print(e)
print("Date mancanti aggiunte")
print("Date")
x=np.array(range(0,len(date)))
hover=HoverTool()
# hover.tooltips = [
#     ("index", "$index"),
#     ("(x,y)", "($x, $y)"),
#     ("radius", "@radius"),
#     ("fill color", "$color[hex, swatch]:fill_color"),
#     ("fill color", "$color[hex]:fill_color"),
#     ("fill color", "$color:fill_color"),
#     ("fill color", "$swatch:fill_color"),
#     ("foo", "@foo"),
#     ("bar", "@bar"),
#     ("baz", "@baz{safe}"),
#     ("total", "@total{$0,0.00}")
# ]
# bokeh create custom tooltip function
from bokeh.plotting import figure, show
from bokeh.layouts import gridplot, Row
from bokeh.models import ColumnDataSource, CDSView, BooleanFilter, CustomJS, BoxSelectTool, HoverTool
def create_string_data(date):
    ris="["
    for d in date:
        ris+=str(d)+","
    ris=ris[:-1]
    ris+="];"
    return ris

code = create_string_data(date)+"""  
var indices = cb_data.index['1d'].indices;
console.log(indices);
if (indices.length > 0){
    if(plot_tooltip.x_range.bounds == null)
    {
        Bokeh.documents[0].add_root(plot_tooltip)
    }
    const idx = indices[0]
    lines.data_source.data['x'] = source.data['xs'][idx]
    lines.data_source.data['y'] = source.data['ys'][idx]
    lines.data_source.change.emit();

    circles.data_source.data['x'] = source.data['xs'][idx]
    circles.data_source.data['y'] = source.data['ys'][idx]
    circles.data_source.change.emit();  

    div = document.getElementsByClassName('bk-root')[1];
    console.log("ciao");
    div.style = "position:absolute; left:" + cb_data.geometry['sx'] + "px; top:" + cb_data.geometry['sy'] + "px;";              
} """


hover = HoverTool()
hover.tooltips = None
p = figure(title="COVID cases", x_axis_label='Dates', y_axis_label='Cases', x_range=(0,len(date)),sizing_mode="stretch_width",height=600,tools=[hover])
p.xaxis.ticker = x
p.xaxis.major_label_orientation = "vertical"
p.xaxis.major_label_overrides = {i: (date[i] if i%30==0 else "") for i in range(len(date))}
print("Genero il grafico")
linee=[]
for y in assi_y:
  valori=[]
  for asse in assi_y.get(y):
    valori.append(asse[0])
  linee.append(p.line(x, valori, legend_label=y, line_width=2))
print("Grafico generato")
show(p)
callback = CustomJS(args = dict(lines=linee), code = code)